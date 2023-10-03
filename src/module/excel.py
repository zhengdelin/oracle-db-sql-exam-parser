import os
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.dimensions import RowDimension
import util.array as array
from module.output import toOutputDir, toRootDir, OUTPUT_DIR
import aspose.cells as asposeCells
from module.zip import zipOutputs
__EXCEL_XLS_FILE_NAME = "output.xls"
__EXCEL_XLSX_FILE_NAME = __EXCEL_XLS_FILE_NAME + "x"


def exportToExcel(results: list):
    toRootDir()
    # excel設置
    SHEET_START_ROW = 3
    CENTER_COLUMNS = [0, 2, 3]  # 置中的欄位
    FONT_STYLE = Font(size=10, color="FF000000", bold=False)
    wb = openpyxl.load_workbook("sample.xlsx")
    sheet = wb.worksheets[0]
    # sheet.column_dimensions['A'].width=10
    # sheet.column_dimensions['B'].width=100
    # sheet.column_dimensions['C'].width=10
    # sheet.column_dimensions['D'].width=10
    # sheet.column_dimensions['E'].width=40

    # 選項
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 25
    sheet.column_dimensions['H'].width = 25
    sheet.column_dimensions['I'].width = 25
    sheet.column_dimensions['J'].width = 25
    sheet.column_dimensions['K'].width = 25
    sheet.column_dimensions['L'].width = 25

    for i in range(0, len(results)):
        item = results[i]
        rowNum = i + 1 + SHEET_START_ROW
        # 將高度重設
        sheet.row_dimensions[rowNum] = RowDimension(
            worksheet=sheet, index=rowNum, height=None)

        itemsWriteToExcel = (item["ansType"], item["subject"], item["difficulty"], item["ans"], item["explanation"]) + \
            tuple(array.map(item["choices"], lambda i: i["text"]))
        for colIndex in range(0, len(itemsWriteToExcel)):

            cell = sheet.cell(rowNum, colIndex + 1)
            isCenterColumn = colIndex in CENTER_COLUMNS
            # 樣式設定
            cell.alignment = Alignment(horizontal=("center" if isCenterColumn else "left"),
                                       wrap_text=True, vertical=("center" if isCenterColumn else "top"))
            cell.font = FONT_STYLE
            cell.value = itemsWriteToExcel[colIndex]

    toOutputDir()
    wb.save(__EXCEL_XLSX_FILE_NAME)

    newWb = asposeCells.Workbook(__EXCEL_XLSX_FILE_NAME)
    os.system(f"del {__EXCEL_XLSX_FILE_NAME}")
    newWb.save(__EXCEL_XLS_FILE_NAME)

    # xlsx to xls
    # workbook = spireXls.Workbook()
    # workbook.LoadFromFile(__EXCEL_XLSX_FILE_NAME)
    # workbook.SaveToFile("to.xls")
    # workbook.Dispose()
    print(
        f"\n成功導出至Excel，已存至 {os.path.join(OUTPUT_DIR, __EXCEL_XLSX_FILE_NAME)}")
    # newWb = openpyxl.Workbook()
    # newWb.copy_worksheet(sheet)
    # newWb.save("outputs/output.xls")
    zipOutputs()
